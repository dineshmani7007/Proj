import streamlit as st
import pandas as pd
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import tempfile
import os

# ================= PAGE CONFIG =================
st.set_page_config(page_title="Fixture Comparison Tool", layout="wide")

st.title("Fixture Comparison Tool")
st.caption("Detect added, removed, modified fixtures (including spelling differences)")

# ================= NORMALIZATION =================
def normalize_desc(text):
    if pd.isna(text):
        return ""
    return " ".join(text.lower().split())

def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()

MATCH_THRESHOLD = 0.9

# ================= STREAMLIT HIGHLIGHT =================
def highlight_row(row):
    if row["Change Type"] == "ADDED":
        return ["background-color: #C6EFCE"] * len(row)
    elif row["Change Type"] == "REMOVED":
        return ["background-color: #FFC7CE"] * len(row)
    elif str(row["Change Type"]).startswith("MODIFIED"):
        return ["background-color: #FFEB9C"] * len(row)
    return [""] * len(row)

# ================= FILE UPLOAD =================
col1, col2 = st.columns(2)

with col1:
    old_file = st.file_uploader("Upload OLD Fixture File", type=["xlsx"])

with col2:
    new_file = st.file_uploader("Upload NEW Fixture File", type=["xlsx"])

if old_file and new_file:
    st.success("Files uploaded successfully!")

    old_df = pd.read_excel(old_file)
    new_df = pd.read_excel(new_file)

    required_cols = ["Start Date", "Start Time", "Description", "Venue"]
    old_df = old_df[required_cols].copy()
    new_df = new_df[required_cols].copy()

    # Normalize descriptions
    old_df["Desc_norm"] = old_df["Description"].apply(normalize_desc)
    new_df["Desc_norm"] = new_df["Description"].apply(normalize_desc)

    # ================= FUZZY MATCHING =================
    matched_new = set()
    pairs = []

    for i, old_row in old_df.iterrows():
        best_match = None
        best_score = 0

        for j, new_row in new_df.iterrows():
            if j in matched_new:
                continue

            score = similarity(old_row["Desc_norm"], new_row["Desc_norm"])
            if score > best_score:
                best_score = score
                best_match = j

        if best_score >= MATCH_THRESHOLD:
            matched_new.add(best_match)
            pairs.append((i, best_match))
        else:
            pairs.append((i, None))

    # Add unmatched NEW rows
    for j in new_df.index:
        if j not in matched_new:
            pairs.append((None, j))

    # ================= BUILD COMPARISON =================
    rows = []

    for old_i, new_i in pairs:
        old = old_df.loc[old_i] if old_i is not None else None
        new = new_df.loc[new_i] if new_i is not None else None

        if old is None:
            change_type = "ADDED"
        elif new is None:
            change_type = "REMOVED"
        else:
            changes = []

            if old["Description"] != new["Description"]:
                changes.append("Description")
            if old["Start Date"] != new["Start Date"]:
                changes.append("Start Date")
            if old["Start Time"] != new["Start Time"]:
                changes.append("Start Time")
            if old["Venue"] != new["Venue"]:
                changes.append("Venue")

            change_type = (
                "MODIFIED (" + ", ".join(changes) + ")"
                if changes else "NO CHANGE"
            )

        rows.append({
            "Change Type": change_type,
            "Description_OLD": old["Description"] if old is not None else "",
            "Description_NEW": new["Description"] if new is not None else "",
            "Start Date_OLD": old["Start Date"] if old is not None else "",
            "Start Time_OLD": old["Start Time"] if old is not None else "",
            "Venue_OLD": old["Venue"] if old is not None else "",
            "Start Date_NEW": new["Start Date"] if new is not None else "",
            "Start Time_NEW": new["Start Time"] if new is not None else "",
            "Venue_NEW": new["Venue"] if new is not None else "",
        })

    final_df = pd.DataFrame(rows)

    # ================= STREAMLIT PREVIEW =================
    st.subheader("Comparison Preview")
    st.dataframe(
        final_df.style.apply(highlight_row, axis=1),
        use_container_width=True
    )

    # ================= EXPORT TO EXCEL =================
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        output_path = tmp.name

    final_df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    FILL_ADDED = PatternFill("solid", fgColor="C6EFCE")
    FILL_REMOVED = PatternFill("solid", fgColor="FFC7CE")
    FILL_MODIFIED = PatternFill("solid", fgColor="FFEB9C")

    header = [cell.value for cell in ws[1]]
    change_col_idx = header.index("Change Type") + 1

    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, change_col_idx).value

        if val == "ADDED":
            fill = FILL_ADDED
        elif val == "REMOVED":
            fill = FILL_REMOVED
        elif str(val).startswith("MODIFIED"):
            fill = FILL_MODIFIED
        else:
            continue

        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).fill = fill

    wb.save(output_path)

    with open(output_path, "rb") as f:
        st.download_button(
            label="Download Highlighted Comparison Report",
            data=f,
            file_name="Fixture_Comparison_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    os.remove(output_path)

else:
    st.info("Please upload both OLD and NEW Excel files to start comparison.")
